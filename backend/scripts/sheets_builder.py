import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def create_excel(file_path, worksheets_info):
    """
    file_path: 要儲存的 Excel 檔案路徑，例如 "test.xlsx"
    worksheets_info: dict，每個 key 是 worksheet 名稱，value 是表頭 list
    """
    wb = Workbook()
    
    # 先刪掉 openpyxl 建立的預設 worksheet
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    
    for ws_name, headers in worksheets_info.items():
        ws = wb.create_sheet(ws_name)
        
        # 寫入表頭
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            # 加粗 + 灰底
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment=Alignment(horizontal='center', vertical='center')
    return wb

def creat_square(ws):
    for r in range(1,51):
        ws.row_dimensions[r].height=15
    for c in range(1,51):
        column_letter=get_column_letter(c)
        ws.column_dimensions[column_letter].width=3

import numpy as np

logo = np.array([
[0,0,0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0,0,0],
[0,0,0,1,2,1,1,1,1,0,0,1,1,1,1,2,1,0,0,0],
[0,0,1,1,1,1,1,2,1,0,0,1,2,1,1,1,1,1,0,0],
[0,0,1,2,2,1,1,2,1,1,1,1,2,1,1,2,2,1,0,0],
[1,1,1,2,1,1,2,2,2,2,2,2,2,2,1,1,2,1,1,1],
[1,2,1,1,1,2,2,2,2,2,2,2,2,2,2,1,1,1,2,1],
[1,2,1,1,2,2,2,2,2,2,2,2,2,2,2,2,1,1,2,1],
[1,2,1,1,2,2,1,1,2,2,2,2,1,1,2,2,1,1,2,1],
[1,2,1,1,2,2,1,1,2,2,2,2,1,1,2,2,1,1,2,1],
[1,1,2,2,2,2,1,1,2,2,2,2,1,1,2,2,2,2,1,1],
[0,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,1,1,1,0],
[0,0,0,1,2,2,2,2,2,2,2,2,2,2,2,2,1,0,0,0],
[0,0,0,1,1,1,2,2,1,1,1,1,2,2,1,1,1,0,0,0],
[0,0,0,0,0,1,2,2,1,0,0,1,2,2,1,0,0,0,0,0],
[0,0,0,1,1,1,1,1,1,0,0,1,1,1,1,1,1,0,0,0],
[0,0,0,1,2,1,0,0,0,0,0,0,0,0,1,2,1,0,0,0],
[0,0,0,1,1,1,0,0,0,0,0,0,0,0,1,1,1,0,0,0]
])


def creat_logo(ws,logo,x_shift,y_shift):
    color_map={
        0:"FFD9D9D9",
        1:"FF808080",
        2:"FFB1A0C7",
    }
    logo=logo
    for i in range(logo.shape[0]):
        for j in range(logo.shape[1]):
            cell=ws.cell(row=i+1+y_shift,column=j+1+x_shift)
            color=color_map.get(logo[i,j])
            cell.fill=PatternFill(start_color=color,end_color=color,fill_type="solid")

def fill_range(ws, start_row, end_row, start_col, end_col, color):
    fill = PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid"
    )
    
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = fill
    return ws

def merge_with_text(ws,up_left,down_right,text,border_control=True,border_color="FF808080",border_style='thick',bg_control=True,bg_color="FFFFFFFF"):
    ws.merge_cells(f'{up_left}:{down_right}')
    ws[up_left]=text
    ws[up_left].alignment=Alignment(wrap_text=True)
    if bg_control==True:
        ws[up_left].fill=PatternFill(start_color=bg_color,end_color=bg_color,fill_type="solid")
    if border_control==True:
        border_style=Side(border_style=border_style,color=border_color)
        start_col, start_row = coordinate_from_string(up_left)
        end_col, end_row = coordinate_from_string(down_right)
        start_col = column_index_from_string(start_col)
        end_col = column_index_from_string(end_col)
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                border_draw = Border(
                    left=border_style if c == start_col else None,
                    right=border_style if c == end_col else None,
                    top=border_style if r == start_row else None,
                    bottom=border_style if r == end_row else None
                )
                ws.cell(row=r, column=c).border = border_draw

def cell_style(ws,up_left,size,font_color,bold_control,horizontal='center',vertical='center'):
    ws[up_left].alignment=Alignment(horizontal=horizontal, vertical=vertical)
    ws[up_left].font = Font(size=size,color=font_color,bold=bold_control)

def set_column_width(ws,header_dic):
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header in header_dic:
            ws.column_dimensions[get_column_letter(col)].width = header_dic[header]


#column width=num_word*3

width_dic_review_reply = {
    "核准並移動": 15,
    "學年": 9
}



# -------------------
# 範例使用
# -------------------
if __name__ == "__main__":
    worksheets_info = {
        "資料庫操作說明與導覽":[],
        "課程評鑑回覆":["核准並移動","學年","母分類","子分類","課程名稱","授課教師","甜度","涼度","有料程度","心得","時間戳記","評鑑ID"],
        "評鑑資料庫": ["課程母分類", "課程子分類", "課程名稱", "授課教師", "修課時間","甜度","涼度","有料程度","評價與修課指引","評鑑ID"],
        "SystemReports": ["Timestamp", "Reporter", "DeviceInfo", "Resolved"],
        "學期課程授課教師對照表": ["修課學期", "課程名稱", "授課教師"],
        "課程資料庫":["課程母分類","課程子分類","課程名稱"],
        "課程分類參考表":["課程母分類","課程子分類"],
        "瀏覽記錄":["課程名稱","授課教師","瀏覽時間","瀏覽次數"],
        "帳號密碼":["帳號","密碼","姓名"],
        "資料庫帳號密碼設定":["帳號","密碼","使用者名稱","最新登入時間","登入次數"]
    }
    wb=create_excel("本地測試.xlsx", worksheets_info)
    ws=wb["資料庫操作說明與導覽"]
    creat_square(ws)
    ws=fill_range(ws,1,21,1,50,"FFD9D9D9")
    creat_logo(ws,logo,3,2)
    merge_with_text(ws,'Z3','AV9',"CGUCMSA\n修課指引與課程評鑑資料庫")
    cell_style(ws,'Z3',22,"FF808080",bold_control=True)
    merge_with_text(ws,'Z11','AD12',"帳號 Account",border_control=False,bg_control=False)
    cell_style(ws,'Z11',12,"FF000000",bold_control=True)
    merge_with_text(ws,'Z15','AD16',"密碼 Password",border_control=False,bg_control=False)
    cell_style(ws,'Z15',12,"FF000000",bold_control=True)
    merge_with_text(ws,'AF11','AM12',"")
    cell_style(ws,'AF11',12,"FF000000",bold_control=False,horizontal='left')
    merge_with_text(ws,'AF15','AM16',"")
    cell_style(ws,'AF15',12,"FF000000",bold_control=False,horizontal='left')
    merge_with_text(ws,'Z19','AV19',"© 長庚大學中醫學系系學會所有",border_control=False,bg_control=False)
    cell_style(ws,'Z19',11,"FF808080",bold_control=True)
    ws=wb["資料庫帳號密碼設定"]
    ws['A2']="default_account"
    ws['B2']="default_password"
    ws['C2']="預設使用者1"
    ws['A3']="test"
    ws['B3']="test"
    ws['C3']="預設使用者2"
    ws=wb["帳號密碼"]
    ws['A2']="default_account"
    ws['B2']="default_password"
    ws['C2']="預設使用者1"
    ws['A3']="test"
    ws['B3']="test"
    ws['C3']="預設使用者2"
    ws=wb["課程評鑑回覆"]
    set_column_width(ws,width_dic_review_reply)
    wb.save("database.xlsx")